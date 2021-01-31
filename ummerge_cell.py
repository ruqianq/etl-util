from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import re
def create_merged_cell_lookup(sheet) -> dict:
    """
    :param sheet:
    :return: the key-value pairs (dict) of merged cell and top value
    e.g {
    'A10:A20': 'Manufacture Of Beer (In Barrels)'
    'B11:B19': 'Removals'
    'C11:C14': 'Taxable*'
    'C15:C18: 'Tax Free'
    }
    """
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        if min_col == max_col:
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            top_left_cell_value = top_left_cell_value.replace('\n', ' ')
            top_left_cell_value = " ".join(re.split("\s+", top_left_cell_value, flags=re.UNICODE))
            merged_lookup[str(cell_group)] = top_left_cell_value
    return merged_lookup
def unmerge_cell_copy_top_value():
    """
    :return: modified work sheet in memory
    """
    wbook = load_workbook("./data/test2.xlsx")
    sheet = wbook["Consolidated"]
    lookup = create_merged_cell_lookup(sheet)
    cell_group_list = lookup.keys()
    for cell_group in cell_group_list:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = lookup[cell_group]
    return sheet
    # wbook.save("openpyxl_merge_unmerge.xlsx")
if __name__ == '__main__':
    unmerge_cell_copy_top_value()
